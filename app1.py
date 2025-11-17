import streamlit as st
import pandas as pd
import re
from io import BytesIO

# Page configuration
st.set_page_config(page_title="RIS Analysis Tool", page_icon="📊", layout="wide")

# Title and description
st.title("📊 RIS (Regional Inventory Storage) Analysis Tool")
st.markdown("Upload your Excel files to analyze RIS vs Non-RIS distribution by brand and state.")

# Helper functions
def normalize(s):
    """Normalize column for joining"""
    return s.astype(str).str.strip().str.upper().replace({"": pd.NA})

def normalize_text(s):
    """Normalize text: replace NBSP, remove punctuation, collapse spaces, lowercase, remove spaces."""
    if pd.isna(s):
        return ""
    s = str(s).replace("\xa0", " ")
    s = s.lower().strip()
    s = re.sub(r"[^\w\s]", "", s)   # remove punctuation
    s = re.sub(r"\s+", " ", s).strip()
    return s.replace(" ", "")

def safe_correct(raw, canon_map):
    """Replace with canonical state name if exact match exists"""
    norm = normalize_text(raw)
    if norm in canon_map:
        return canon_map[norm]
    return raw

# File uploaders
st.sidebar.header("📁 Upload Files")
original_file = st.sidebar.file_uploader("Upload Original.xlsx", type=['xlsx'])
fc_file = st.sidebar.file_uploader("Upload FC Stat.xlsx", type=['xlsx'])
pm_file = st.sidebar.file_uploader("Upload PM.xlsx", type=['xlsx'])

if original_file and fc_file and pm_file:
    try:
        # Load data
        with st.spinner("Loading data..."):
            Working = pd.read_excel(original_file)
            FC = pd.read_excel(fc_file)
            PM = pd.read_excel(pm_file)
        
        st.success("✅ All files loaded successfully!")
        
        # Display file info
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Original Records", len(Working))
        with col2:
            st.metric("FC Records", len(FC))
        with col3:
            st.metric("PM Records", len(PM))
        
        # Process button
        if st.button("🔄 Process Data", type="primary"):
            with st.spinner("Processing data..."):
                
                # Step 1: Merge FC data
                key_col_name = "fulfillment-center-id"
                required_cols = ["FC", "State", "Cluster"]
                
                if key_col_name not in Working.columns:
                    st.error(f"Key column '{key_col_name}' missing in Working file.")
                    st.stop()
                
                missing = [c for c in required_cols if c not in FC.columns]
                if missing:
                    st.error(f"FC file missing columns: {missing}")
                    st.stop()
                
                # Create FC lookup
                FC_lookup = FC[["FC", "State", "Cluster"]].copy()
                FC_lookup.columns = ["lookup_key", "fulfillment_state", "cluster"]
                
                # Create temporary join keys
                Working["_join"] = normalize(Working[key_col_name])
                FC_lookup["_join"] = normalize(FC_lookup["lookup_key"])
                
                # Merge
                Working = Working.merge(
                    FC_lookup[["_join", "fulfillment_state", "cluster"]],
                    how="left",
                    left_on="_join",
                    right_on="_join"
                )
                
                # Drop helper columns
                Working.drop(columns=["_join"], inplace=True)
                
                # Step 2: Normalize ship-state
                fc_states = FC["State"].dropna().astype(str).str.strip().tolist()
                canon_map = { 
                    normalize_text(s): s.strip() 
                    for s in fc_states
                    if normalize_text(s) != ""
                }
                
                # Backup original
                if "ship-state_original" not in Working.columns:
                    Working["ship-state_original"] = Working["ship-state"]
                
                # Correct state names
                Working["ship-state"] = Working["ship-state_original"].apply(
                    lambda x: safe_correct(x, canon_map)
                )
                
                # Step 3: Create RIS Status
                Working["RIS Status"] = Working.apply(
                    lambda row: "RIS" if str(row["ship-state_original"]).strip().replace(" ", "") ==
                                          str(row["fulfillment_state"]).strip().replace(" ", "")
                                else "Non RIS",
                    axis=1
                )
                
                # Step 4: Merge PM data (Brand)
                PM_lookup = PM.iloc[:, 2:7].copy()
                PM_lookup.columns = ["lookup_key", "col2", "col3", "col4", "Brand"]
                
                # Assuming 'sku' is the key column - adjust if needed
                lookup_key_col = "sku"
                if lookup_key_col in Working.columns:
                    Working = Working.merge(
                        PM_lookup[["lookup_key", "Brand"]],
                        how="left",
                        left_on=lookup_key_col,
                        right_on="lookup_key"
                    )
                    Working.drop(columns=["lookup_key"], inplace=True)
                
                # Step 5: Create detailed pivot table with subtotals
                # First create the main pivot
                detailed_pivot = pd.pivot_table(
                    Working,
                    values="quantity-shipped",
                    index=["Brand", "fulfillment_state", "ship-state"],
                    columns="RIS Status",
                    aggfunc="sum",
                    fill_value=0
                )
                
                # Add Grand Total column
                detailed_pivot["Grand Total"] = detailed_pivot.sum(axis=1)
                
                # Create brand totals
                brand_totals = detailed_pivot.groupby(level=0).sum()
                brand_totals.index = [(brand, f"{brand} Total", "") for brand in brand_totals.index]
                
                # Create fulfillment state totals
                state_totals = detailed_pivot.groupby(level=[0, 1]).sum()
                state_totals.index = [(brand, f"{state} Total", "") for brand, state in state_totals.index]
                
                # Combine all data
                detailed_pivot_with_totals = pd.concat([detailed_pivot, state_totals, brand_totals])
                detailed_pivot_with_totals = detailed_pivot_with_totals.sort_index()
                
                # Add overall grand total
                grand_total = detailed_pivot.sum()
                grand_total.name = ("Grand Total", "", "")
                detailed_pivot_with_totals = pd.concat([
                    detailed_pivot_with_totals, 
                    pd.DataFrame([grand_total], index=[("Grand Total", "", "")])
                ])
                
                # Step 6: Amazon Inventory Placement Program Analysis
                # Local FC mapping based on table
                LOCAL_MAP = {
                    "DED3": ["DEL4", "DEL5", "DED4"],
                    "DED5": ["DEL4", "DEL5", "DED4"],
                    "ISK3": ["BOM5", "BOM7", "PNQ3"],
                    "BLR4": ["BLR7", "BLR8"]
                }
                
                def ris_status_from_ixd(row):
                    rc = row.get("Inferred_Receive_Centre")
                    fc = str(row.get("fulfillment-center-id", "")).strip().upper()
                    if pd.isna(rc) or rc not in LOCAL_MAP:
                        return "Non RIS"
                    # If fulfillment center in LOCAL list → RIS
                    if fc in LOCAL_MAP[rc]:
                        return "RIS"
                    else:
                        return "Non RIS"
                
                Working["RIS_by_Table"] = Working.apply(ris_status_from_ixd, axis=1)
                
                # Create inventory placement pivot with subtotals
                inventory_pivot = pd.pivot_table(
                    Working,
                    values="quantity-shipped",
                    index=["Brand", "fulfillment_state", "ship-state"],
                    columns="RIS_by_Table",
                    aggfunc="sum",
                    fill_value=0
                )
                
                # Ensure both RIS and Non RIS columns exist
                if "RIS" not in inventory_pivot.columns:
                    inventory_pivot["RIS"] = 0
                if "Non RIS" not in inventory_pivot.columns:
                    inventory_pivot["Non RIS"] = 0
                
                # Reorder columns to have Non RIS, RIS
                column_order = ["Non RIS", "RIS"]
                inventory_pivot = inventory_pivot[[col for col in column_order if col in inventory_pivot.columns]]
                
                # Add Grand Total column
                inventory_pivot["Grand Total"] = inventory_pivot.sum(axis=1)
                
                # Create brand totals
                inv_brand_totals = inventory_pivot.groupby(level=0).sum()
                inv_brand_totals.index = [(brand, f"{brand} Total", "") for brand in inv_brand_totals.index]
                
                # Create fulfillment state totals
                inv_state_totals = inventory_pivot.groupby(level=[0, 1]).sum()
                inv_state_totals.index = [(brand, f"{state} Total", "") for brand, state in inv_state_totals.index]
                
                # Combine all data
                inventory_pivot_with_totals = pd.concat([inventory_pivot, inv_state_totals, inv_brand_totals])
                inventory_pivot_with_totals = inventory_pivot_with_totals.sort_index()
                
                # Add overall grand total
                inv_grand_total = inventory_pivot.sum()
                inv_grand_total.name = ("Grand Total", "", "")
                inventory_pivot_with_totals = pd.concat([
                    inventory_pivot_with_totals, 
                    pd.DataFrame([inv_grand_total], index=[("Grand Total", "", "")])
                ])
                
                # Brand-level summary for inventory placement
                inventory_brand_summary = pd.pivot_table(
                    Working,
                    values="quantity-shipped",
                    index=["Brand"],
                    columns="RIS_by_Table",
                    aggfunc="sum",
                    fill_value=0,
                    margins=True,
                    margins_name="Grand Total"
                )
                
                # Ensure both columns exist before calculating percentages
                if "RIS" not in inventory_brand_summary.columns:
                    inventory_brand_summary["RIS"] = 0
                if "Non RIS" not in inventory_brand_summary.columns:
                    inventory_brand_summary["Non RIS"] = 0
                
                # Reorder columns
                column_order = ["Non RIS", "RIS", "Grand Total"]
                inventory_brand_summary = inventory_brand_summary[[col for col in column_order if col in inventory_brand_summary.columns]]
                
                # Calculate Grand Total if it doesn't exist
                if "Grand Total" not in inventory_brand_summary.columns:
                    inventory_brand_summary["Grand Total"] = inventory_brand_summary.sum(axis=1)
                
                # Add percentage columns
                inventory_brand_summary["Non RIS%"] = inventory_brand_summary["Non RIS"] / inventory_brand_summary["Grand Total"]
                inventory_brand_summary["RIS%"] = inventory_brand_summary["RIS"] / inventory_brand_summary["Grand Total"]
                inventory_brand_summary[["Non RIS%", "RIS%"]] = inventory_brand_summary[["Non RIS%", "RIS%"]].fillna(0)
                
                # Brand-level summary
                brand_summary = pd.pivot_table(
                    Working,
                    values="quantity-shipped",
                    index=["Brand"],
                    columns="RIS Status",
                    aggfunc="sum",
                    fill_value=0,
                    margins=True,
                    margins_name="Grand Total"
                )
                brand_summary["Non RIS%"] = brand_summary["Non RIS"] / brand_summary["Grand Total"]
                brand_summary["RIS%"] = brand_summary["RIS"] / brand_summary["Grand Total"]
                brand_summary[["Non RIS%", "RIS%"]] = brand_summary[["Non RIS%", "RIS%"]].fillna(0)
                
            st.success("✅ Processing complete!")
            
            # Create tabs for different analyses
            tab1, tab2 = st.tabs(["📊 State-Based RIS Analysis", "🏭 Amazon Inventory Placement Program"])
            
            with tab1:
                st.header("📈 State-Based RIS Analysis")
                
                # Brand Summary
                st.subheader("Brand-Level Summary")
                st.dataframe(
                    brand_summary.style.format({
                        "Non RIS": "{:,.0f}",
                        "RIS": "{:,.0f}",
                        "Grand Total": "{:,.0f}",
                        "Non RIS%": "{:.2%}",
                        "RIS%": "{:.2%}"
                    }),
                    use_container_width=True
                )
                
                # Detailed pivot
                st.subheader("Detailed Analysis (Brand → Fulfillment State → Ship State)")
                
                # Format the dataframe for display
                def format_pivot_display(df):
                    """Format the pivot table for better display"""
                    display_df = df.copy()
                    
                    # Create display index
                    display_index = []
                    for idx in display_df.index:
                        brand, fstate, sstate = idx
                        if "Total" in fstate:
                            display_index.append((f"  {brand}", "", ""))
                        elif sstate == "":
                            display_index.append((brand, fstate, ""))
                        else:
                            display_index.append((brand, fstate, sstate))
                    
                    display_df.index = pd.MultiIndex.from_tuples(
                        display_index, 
                        names=["Brand", "Fulfillment State", "Ship State"]
                    )
                    
                    return display_df
                
                formatted_pivot = format_pivot_display(detailed_pivot_with_totals)
                
                # Simple styling without set_properties to avoid index conflicts
                st.dataframe(
                    formatted_pivot.style.format({
                        "Non RIS": "{:,.0f}",
                        "RIS": "{:,.0f}",
                        "Grand Total": "{:,.0f}"
                    }),
                    use_container_width=True,
                    height=600
                )
            
            with tab2:
                st.header("🏭 Amazon Inventory Placement Program Analysis")
                st.info("📌 Based on Inferred Receive Centre to Fulfillment Center mapping")
                
                # Show mapping
                with st.expander("🗺️ View Local FC Mapping"):
                    st.code("""
LOCAL_MAP = {
    "DED3": ["DEL4", "DEL5", "DED4"],
    "DED5": ["DEL4", "DEL5", "DED4"],
    "ISK3": ["BOM5", "BOM7", "PNQ3"],
    "BLR4": ["BLR7", "BLR8"]
}
                    """)
                
                # Brand Summary for Inventory Placement
                st.subheader("Brand-Level Summary (Inventory Placement)")
                st.dataframe(
                    inventory_brand_summary.style.format({
                        "Non RIS": "{:,.0f}",
                        "RIS": "{:,.0f}",
                        "Grand Total": "{:,.0f}",
                        "Non RIS%": "{:.2%}",
                        "RIS%": "{:.2%}"
                    }),
                    use_container_width=True
                )
                
                # Detailed Inventory Placement Pivot
                st.subheader("Detailed Analysis (Brand → Fulfillment State → Ship State)")
                
                formatted_inv_pivot = format_pivot_display(inventory_pivot_with_totals)
                
                # Simple styling without set_properties to avoid index conflicts
                st.dataframe(
                    formatted_inv_pivot.style.format({
                        "Non RIS": "{:,.0f}",
                        "RIS": "{:,.0f}",
                        "Grand Total": "{:,.0f}"
                    }),
                    use_container_width=True,
                    height=600
                )
            
            # Download buttons
            st.subheader("📥 Download Results")
            
            col1, col2, col3, col4 = st.columns(4)
            
            # Download processed data
            with col1:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    Working.to_excel(writer, sheet_name='Processed Data', index=False)
                output.seek(0)
                st.download_button(
                    label="📄 Processed Data",
                    data=output,
                    file_name="processed_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Download state-based analysis
            with col2:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    brand_summary.to_excel(writer, sheet_name='Brand Summary')
                    detailed_pivot_with_totals.to_excel(writer, sheet_name='Detailed Analysis')
                    
                    # Format both sheets
                    from openpyxl.styles import Font, PatternFill
                    
                    for sheet_name in ['Brand Summary', 'Detailed Analysis']:
                        worksheet = writer.sheets[sheet_name]
                        for row in worksheet.iter_rows(min_row=2):
                            if row[1].value and "Total" in str(row[1].value):
                                for cell in row:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                output.seek(0)
                st.download_button(
                    label="📊 State-Based RIS",
                    data=output,
                    file_name="state_based_ris_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Download inventory placement analysis
            with col3:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    inventory_brand_summary.to_excel(writer, sheet_name='Brand Summary')
                    inventory_pivot_with_totals.to_excel(writer, sheet_name='Detailed Analysis')
                    
                    # Format both sheets
                    from openpyxl.styles import Font, PatternFill
                    
                    for sheet_name in ['Brand Summary', 'Detailed Analysis']:
                        worksheet = writer.sheets[sheet_name]
                        for row in worksheet.iter_rows(min_row=2):
                            if row[1].value and "Total" in str(row[1].value):
                                for cell in row:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                output.seek(0)
                st.download_button(
                    label="🏭 Inventory Placement",
                    data=output,
                    file_name="inventory_placement_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Download combined report
            with col4:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # State-based analysis
                    brand_summary.to_excel(writer, sheet_name='State-Based Brand Summary')
                    detailed_pivot_with_totals.to_excel(writer, sheet_name='State-Based Detailed')
                    
                    # Inventory placement analysis
                    inventory_brand_summary.to_excel(writer, sheet_name='Inventory Brand Summary')
                    inventory_pivot_with_totals.to_excel(writer, sheet_name='Inventory Detailed')
                    
                    # Raw processed data
                    Working.to_excel(writer, sheet_name='Processed Data', index=False)
                    
                    # Format all sheets
                    from openpyxl.styles import Font, PatternFill
                    
                    for sheet_name in writer.sheets:
                        if sheet_name != 'Processed Data':
                            worksheet = writer.sheets[sheet_name]
                            for row in worksheet.iter_rows(min_row=2):
                                if row[1].value and "Total" in str(row[1].value):
                                    for cell in row:
                                        cell.font = Font(bold=True)
                                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                output.seek(0)
                st.download_button(
                    label="📦 Complete Report",
                    data=output,
                    file_name="complete_ris_analysis_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.exception(e)

else:
    st.info("👈 Please upload all three required Excel files to begin analysis.")
    
    with st.expander("ℹ️ File Requirements"):
        st.markdown("""
        ### Required Files:
        
        1. **Original.xlsx** - Main order data containing:
           - `fulfillment-center-id`: FC code
           - `ship-state`: Delivery state
           - `quantity-shipped`: Quantity
           - `sku`: Product SKU
        
        2. **FC Stat.xlsx** - Fulfillment center data containing:
           - `FC`: FC code
           - `State`: FC state
           - `Cluster`: FC cluster
        
        3. **PM.xlsx** - Product master data containing:
           - Brand information (column G)
           - SKU mapping
        
        ### What this tool does:
        - Matches orders with FC locations
        - Determines RIS (same state) vs Non-RIS (different state) shipments
        - Generates brand-level and detailed state-level analysis
        - Calculates percentages and totals
        """)

# Footer
st.markdown("---")
st.markdown("*RIS Analysis Tool v1.0 - Analyze regional inventory storage patterns*")